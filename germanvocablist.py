import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import re

# Constants for colors
ARTICLE_CONV = {
    "m": "der",  
    "f": "die",  
    "nt": "das", 
    "": "",
}

ARTICLE_COLORS = {
    "der": "0000FF",  # Blue
    "die": "FF0000",  # Red
    "das": "008000",  # Green
    "": "8B4513",     # Brown for no article
    "verb": "800080"  # Purple for verbs
}

TERMINAL_COLORS = {
    "der": "\033[1;34m",  # Blue
    "die": "\033[1;31m",  # Red
    "das": "\033[1;32m",  # Green
    "": "\033[38;2;165;42;42m"      # Brown for no article
}

YELLOW = "\033[1;33m"
MAGENTA = "\033[1;35m"
RESET = "\033[0m"

EXCEL_FILE = "German Words.xlsx"


def get_word_data(word):
    """Fetches the article and definitions of a word from PONS."""
    url = f"https://en.pons.com/translate/german-english/{word.lower()}"
    response = requests.get(url)
    
    if response.status_code != 200:
        print("Error: Unable to fetch data from PONS.")
        return None, None

    soup = BeautifulSoup(response.text, "html.parser")

    # Try to extract the article and definitions
    try:
        # Extract word classes and articles
        wordclass_tags = soup.find_all("span", class_="wordclass")
        word_classes = []
        for tag in wordclass_tags:
            word_class = tag.text.strip()
            article_tag = tag.find_next_sibling("span", class_="genus")
            article = article_tag.text.strip() if article_tag else ""
            word_classes.append((word_class, article))
        word_classes = list(dict.fromkeys(word_classes))  # Remove duplicates

        is_verb = False
        if len(word_classes) > 1 and any(wc in ["N", "VB"] for wc, _ in word_classes):
            print("Multiple word classes found:")
            print(f"{YELLOW}1.{RESET} Noun")
            print(f"{YELLOW}2.{RESET} Verb")
            print(f"{YELLOW}3.{RESET} Other")
            while True:
                try:
                    choice = int(input("Select the appropriate word class (enter the number): ").strip())
                    if choice == 1:
                        article = next((a for wc, a in word_classes if wc == "N"), "")
                        break
                    elif choice == 2:
                        article = ""
                        is_verb = True
                        break
                    elif choice == 3:
                        article = ""
                        break
                    else:
                        print("Invalid choice: Please enter 1, 2, or 3.")
                except ValueError:
                    print("Invalid input: Please enter a number for the word class choice.")
        elif any(wc in ["VB"] for wc, _ in word_classes):
            is_verb = True


        # Exclude "seealso" sections
        seealso_sections = soup.find_all("div", class_="seealso")
        for section in seealso_sections:
            section.decompose()  # Remove the seealso section

        definition_tags = soup.find_all("div", class_="translations")
        definitions = []
        for tag in definition_tags:
            target_tags = tag.find_all("div", class_="target", limit=2)
            for target in target_tags:
                # Remove span elements with the "info" class
                for span in target.find_all("span", class_="info"):
                    span.decompose()
                definition = " ".join(target.stripped_strings)
                definitions.append(definition)
                
        definitions = list(dict.fromkeys(definitions))  # Remove duplicates

        if not definitions:
            return article, "", is_verb

        # Allow user to select the appropriate definition
        print("Choose the appropriate definition:")
        index = 0
        while True:
            for i, definition in enumerate(definitions[index:index+10], start=index+1):
                print(f"{YELLOW}{i}.{RESET} {definition}")
            print(f"{MAGENTA}a.{RESET} Add your own definition")
            print(f"{MAGENTA}m.{RESET} Show more definitions")

            choice = input("Select the appropriate definition (enter the number or key): ").strip().lower()
            if choice.isdigit():
                choice = int(choice)
                if 0 < choice <= len(definitions):
                    selected_definition = definitions[choice - 1]
                    break
                else:
                    print("Invalid choice: Please enter a valid number.")
            elif choice == 'a':
                selected_definition = input("Enter your custom definition: ").strip()
                break
            elif choice == 'm':
                index += 10
                if index >= len(definitions):
                    print("No more definitions available.")
                    index = 0
            else:
                print("Invalid input: Please enter a valid number or key.")

        return article, selected_definition, is_verb
    except AttributeError:
        print("Warning: Could not determine the article or definitions.")
        return "", "", False


def get_verb_conjugations(verb):
    """Fetches the conjugations for Indikativ Präsens from PONS."""
    url = f"https://en.pons.com/verb-tables/german/{verb.lower()}"
    response = requests.get(url)
    
    if response.status_code != 200:
        print("Error: Unable to fetch conjugations from PONS.")
        return None

    soup = BeautifulSoup(response.text, "html.parser")
    conjugations = {}

    try:
        table = soup.find("table", class_="table")
        rows = table.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            if len(cols) == 2:
                person = cols[0].text.strip()
                conjugation = cols[1].text.strip()
                if "ich" in person:
                    conjugations["ich"] = conjugation
                elif "du" in person:
                    conjugations["du"] = conjugation
                elif "er/sie/es" in person:
                    conjugations["er/sie/es"] = conjugation
                elif "wir" in person:
                    conjugations["wir"] = conjugation
                elif "ihr" in person:
                    conjugations["ihr"] = conjugation
                elif "sie" in person:
                    conjugations["sie/Sie"] = conjugation
        return conjugations
    except AttributeError:
        print("Warning: Could not determine the conjugations.")
        return None


def create_or_load_excel():
    """Creates a new Excel file if it doesn't exist, otherwise loads it."""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "General"
        ws.append(["Word", "Definition"])
        bold_font = Font(bold=True, size=16)
        alignment_center = Alignment(horizontal="center", vertical="center")
        for col in range(1, 3):
            ws.cell(row=1, column=col).font = bold_font
            ws.cell(row=1, column=col).alignment = alignment_center

        # Create sheets for each article
        for article in ["der", "die", "das", "No Article"]:
            ws = wb.create_sheet(title=article)
            ws.append(["Word", "Definition"])
            for col in range(1, 3):
                ws.cell(row=1, column=col).font = bold_font
                ws.cell(row=1, column=col).alignment = alignment_center

        wb.save(EXCEL_FILE)
    return wb


def check_duplicate(word, definition, wb):
    """Checks if a word already exists in the Excel file."""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0].lower() == word.lower() and row[1].lower() == definition.lower():
                return row[0], row[1]  # Return word and definition
    return None, None


def create_lesson_sheet(wb, lesson):
    """Creates a new lesson sheet if it doesn't exist."""
    lesson_sheet_name = f"A1.1 - L{lesson}"
    if lesson_sheet_name not in wb.sheetnames:
        lesson_ws = wb.create_sheet(title=lesson_sheet_name)
        lesson_ws.append(["Word", "Definition"])
        bold_font = Font(bold=True, size=16)
        alignment_center = Alignment(horizontal="center", vertical="center")
        for col in range(1, 3):
            lesson_ws.cell(row=1, column=col).font = bold_font
            lesson_ws.cell(row=1, column=col).alignment = alignment_center
    else:
        lesson_ws = wb[lesson_sheet_name]
    return lesson_ws


def create_verbs_sheet(wb):
    """Creates a new verbs sheet if it doesn't exist."""
    if "Verbs" not in wb.sheetnames:
        verbs_ws = wb.create_sheet(title="Verbs")
        verbs_ws.append(["Verb", "Definition", "ich", "du", "er/sie/es", "wir", "ihr", "sie/Sie"])
        bold_font = Font(bold=True, size=14)
        alignment_center = Alignment(horizontal="center", vertical="center")
        for col in range(1, 9):
            verbs_ws.cell(row=1, column=col).font = bold_font
            verbs_ws.cell(row=1, column=col).alignment = alignment_center
    else:
        verbs_ws = wb["Verbs"]
    return verbs_ws


def sort_and_color_sheet(ws):
    """Sorts the sheet alphabetically by the word itself, ignoring the articles, and optionally changes the colors."""
    data = list(ws.iter_rows(min_row=2, values_only=True))
    sorted_data = sorted(data, key=lambda x: x[0].split(" ", 1)[1].lower() if " " in x[0] else x[0].lower())
    for idx, row in enumerate(sorted_data, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col, value=value)
            if col == 1:  # Apply color to the word cell
                checked_article = value.split(" ", 1)[0] if " " in value else ""
                color = ARTICLE_COLORS.get(checked_article, ARTICLE_COLORS[""])
                ws.cell(row=idx, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.cell(row=idx, column=col).font = Font(color="FFFFFF", bold=True)


def add_word_to_sheet(ws, full_word, definition):
    """Adds a word to a specific sheet and sorts it."""
    ws.append([full_word, definition])
    sort_and_color_sheet(ws)


def add_verb_to_sheet(ws, verb, definition, conjugations):
    """Adds a verb to the verbs sheet with its conjugations."""
    ws.append([verb, definition, conjugations.get("ich"), conjugations.get("du"), conjugations.get("er/sie/es"), conjugations.get("wir"), conjugations.get("ihr"), conjugations.get("sie/Sie")])
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        for col, value in enumerate(row, start=1):
            if col == 1:  # Apply purple color to the verb cell
                ws.cell(row=idx, column=col).fill = PatternFill(start_color=ARTICLE_COLORS["verb"], end_color=ARTICLE_COLORS["verb"], fill_type="solid")
                ws.cell(row=idx, column=col).font = Font(color="FFFFFF", bold=True)


def add_word_to_excel(word, article, definition, lesson, wb):
    """Adds a word to the general sheet, the relevant article sheet, and the relevant lesson sheet in the Excel file."""  
    
    # Combine the article and word
    real_article = ARTICLE_CONV.get(article)

    # Capitalize the word based on its type
    if real_article:
        word = word.capitalize()
    else:
        word = word.lower()
        
    full_word = f"{real_article} {word}".strip()

    # Color the terminal output based on the article
    terminal_color = TERMINAL_COLORS.get(real_article, RESET)

    # Check for duplicates in the general sheet
    existing_word, existing_definition = check_duplicate(full_word, definition, wb)
    if existing_word:        
        print(f"⚠️ The word already exists! {terminal_color}{existing_word}{RESET} : {existing_definition}")
        return

    # Add the word to the relevant article sheet if it's not a verb
    if not is_verb:
        general_ws = wb["General"]
        add_word_to_sheet(general_ws, full_word, definition)

        article_sheet_name = real_article if real_article else "No Article"
        article_ws = wb[article_sheet_name]
        add_word_to_sheet(article_ws, full_word, definition)

    # Add the word to the relevant lesson sheet
    lesson_ws = create_lesson_sheet(wb, lesson)
    add_word_to_sheet(lesson_ws, full_word, definition)

    # If the word is a verb, add it to the verbs sheet with conjugations
    if is_verb:
        verbs_ws = create_verbs_sheet(wb)
        conjugations = get_verb_conjugations(word)
        if conjugations:
            add_verb_to_sheet(verbs_ws, word, definition, conjugations)

    wb.save(EXCEL_FILE)
    print(f"✅ Added '{terminal_color}{full_word}{RESET}' : {definition}.")


if __name__ == "__main__":
    wb = create_or_load_excel()

    while True:
        word = input("Enter the word (or 'q' to quit): ").strip()
        if word == "q":
            break

        article, definition, is_verb = get_word_data(word)
        if not definition:
            print("Skipping word due to missing data.")
            continue

        while True:
            try:
                lesson = int(input("Which lesson did you learn this word in? ").strip())
                break
            except ValueError:
                print("Invalid input: Please enter a number for the lesson.")
        
        add_word_to_excel(word, article, definition, lesson, wb)
